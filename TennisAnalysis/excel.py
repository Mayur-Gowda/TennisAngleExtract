import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import column_index_from_string as cl, get_column_letter


class ExcelSave:
    """Class specifically created for saving joint angles in an Excel sheet.
        Saves Frame of Force in one sheet, Frame of Stability in another sheet."""

    def __init__(self, filename, side):
        self.workbook = None
        self.side = side
        self.fileName = os.path.join(
            os.path.dirname(__file__), '../AnalyzedAngles/ExcelSheets', f"{filename}.xlsx")
        self.checkFile()

    @staticmethod
    def create_sheet(sheet, title, **kwargs):
        """Create Excel sheets with headers and borders to better understanding and visualization"""

        sheet.title = title
        sheet['B2'] = kwargs["header"]

        # Merge cells from B2 to V3
        sheet.merge_cells('B2:V3')

        # Apply alignment to the merged cells
        sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')

        # Add top and bottom borders to the merged cell
        for row in sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=22):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

        # Row 4 and 5 are left empty
        # Row 6: C6, H6, M6, R6 are named to the following values: FrontView, TopView, SideView, and NormView
        cols = ['C6', 'H6', 'M6', 'R6']
        colsName = ["FrontView", "TopView", "SideView", "NormView"]

        for col, colName in zip(cols, colsName):
            sheet[col] = colName

        # Merge cells from C6 to G6 and H6 to L6 and M6 to Q6 and R6 to V6
        cellMerge = ['G6', 'L6', 'Q6', 'V6']
        for col, mergeCell in zip(cols, cellMerge):
            sheet.merge_cells(f'{col}:{mergeCell}')

        for col in cols:
            sheet[col].alignment = Alignment(horizontal='center', vertical='center')

        # Add thin borders to the merged cells
        for row in [6]:
            for col_range in ['C', 'H', 'M', 'R']:
                for col in range(sheet[col_range + '6'].column, sheet[col_range + '6'].column + 5):
                    sheet.cell(row=row, column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

        # Row 7 is left empty
        # Row 8: B8, C8, D8, E8, F8, G8 are named to Picture, LeftElbow, LeftShoulder, UpRightHip, DownLeftHip, LeftKnee
        labels = ["Picture", "Elbow", "Shoulder", "UpHip", "DownHip", "Knee"]

        values = {
            "Left": "Right",
            "Up": "Down",
        }

        values = {**values, **{v: k for k, v in values.items()}}

        for col_index, label in enumerate(labels):
            # Write label only once for Picture
            if col_index == 0:
                sheet.cell(row=8, column=2).value = label
                sheet.cell(row=8, column=2).border = Border(top=Side(style='thick'),
                                                            bottom=Side(style='thick'),
                                                            right=Side(style='thick'),
                                                            left=Side(style='thick'))
            else:
                if col_index > 2:
                    sheet.cell(row=8, column=2 + col_index).value = f"{values[kwargs['side']]}{label}"
                else:
                    sheet.cell(row=8, column=2 + col_index).value = f"{kwargs['side']}{label}"
                sheet.cell(row=8, column=2 + col_index).border = Border(top=Side(style='thick'),
                                                                        bottom=Side(style='thick'))

                if col_index == 5:
                    sheet.cell(row=8, column=2 + col_index).border = Border(top=Side(style='thick'),
                                                                            bottom=Side(style='thick'),
                                                                            right=Side(style='thick'))

                # Repeat the other labels without space
                if col_index != 0:
                    for i in range(1, 4):
                        if col_index > 2:
                            sheet.cell(row=8, column=2 + 5 * i + col_index).value = f"{values[kwargs['side']]}{label}"
                        else:
                            sheet.cell(row=8, column=2 + 5 * i + col_index).value = f"{kwargs['side']}{label}"
                        sheet.cell(row=8, column=2 + 5 * i + col_index).border = Border(top=Side(style='thick'),
                                                                                        bottom=Side(style='thick'))
                        if col_index == 5:
                            sheet.cell(row=8, column=2 + 5 * i + col_index).border = Border(top=Side(style='thick'),
                                                                                            bottom=Side(style='thick'),
                                                                                            right=Side(style='thick'))

        # Set column widths and row heights for better spacing
        for col in range(2, 23):
            sheet.column_dimensions[chr(64 + col)].width = 14

    def create_excel_sheets(self):
        """Create the two Frame sheets"""
        if self.side == "left":
            sheet1 = self.workbook.active
            self.create_sheet(sheet1, "LHFOF", header="Frame Of Force", side="Left")
            sheet2 = self.workbook.create_sheet()
            self.create_sheet(sheet2, "LHFOS", header="Frame Of Stability", side="Right")
        elif self.side == "right":
            sheet1 = self.workbook.active
            self.create_sheet(sheet1, "RHFOF", header="Frame Of Force", side="Right")
            sheet2 = self.workbook.create_sheet()
            self.create_sheet(sheet2, "RHFOS", header="Frame Of Stability", side="Left")
        self.workbook.save(self.fileName)

    def addValues(self, sheet_no, df):
        """Add the angle values row by row, by taking the Image name as the identifier."""

        sheet = self.workbook.worksheets[sheet_no]
        row = sheet.max_row + 1
        for index, df_row in df.iterrows():
            sheet.cell(row=row, column=2).value = df_row.name
            for i, val in enumerate(df_row, start=3):
                sheet.cell(row=row, column=i).value = val
            sheet.cell(row=row, column=2).border = Border(left=Side(style='thick'), right=Side(style='thick'))
            for i in range(1, 5):
                sheet.cell(row=row, column=2 + 5 * i).border = Border(right=Side(style='thick'))
            row += 1

        self.workbook.save(self.fileName)

    def checkFile(self):
        """Check if the Excel file is available. If not, it creates a new one."""

        try:
            self.workbook = load_workbook(self.fileName)
        except FileNotFoundError:
            self.workbook = Workbook()
            self.create_excel_sheets()


def sheetSetup(frame, filename, side, df):
    side = side.title()
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        print("Error: Excel sheet not found")
        return
    if frame == "force":
        sheet = wb.create_sheet("P-FOF")
    else:
        sheet = wb.create_sheet("P-FOS")

    sheet['B2'] = f"P-Frame of {frame.title()}"
    sheet.merge_cells('B2:BJ3')
    sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=62):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    cols = ['C6', 'R6', 'AG6', 'AV6']
    cellMerge = ['Q6', 'AF6', 'AU6', 'BJ6']
    colsName = ["FrontView", "TopView", "SideView", "NormView"]
    for col, colName in zip(cols, colsName):
        sheet[col] = colName

    for col, mergeCell in zip(cols, cellMerge):
        sheet.merge_cells(f'{col}:{mergeCell}')
        sheet[col].alignment = Alignment(horizontal='center', vertical='center')

    for col_range in ['C', 'R', 'AG', 'AV']:
        for col in range(sheet[col_range + '6'].column, sheet[col_range + '6'].column + 15):
            sheet.cell(row=6, column=col).border = Border(top=Side(style='thin'),
                                                          bottom=Side(style='thin'),
                                                          left=Side(style='thin'),
                                                          right=Side(style='thin'))

    labels = ["Picture", "Elbow", "Shoulder", "UpHip", "DownHip", "Knee"]

    values = {
        "Left": "Right",
        "Up": "Down",
    }

    values = {**values, **{v: k for k, v in values.items()}}

    j = 0
    for col_index, label in enumerate(labels):
        # Write label only once for Picture

        if col_index == 0:
            sheet.cell(row=9, column=2).value = label
            sheet.cell(row=9, column=2).border = Border(top=Side(style='thin'),
                                                        bottom=Side(style='thin'),
                                                        right=Side(style='thin'),
                                                        left=Side(style='thin'))
        else:
            if col_index > 2:
                sheet.cell(row=8, column=2 + col_index + j).value = f"{values[side]}{label}"
            else:
                sheet.cell(row=8, column=2 + col_index + j).value = f"{side}{label}"

            for i in range(1, 4):
                if col_index > 2:
                    sheet.cell(row=8, column=2 + 15 * i + col_index + j).value = f"{values[side]}{label}"
                else:
                    sheet.cell(row=8, column=2 + 15 * i + col_index + j).value = f"{side}{label}"

            j += 2

    # Convert column letters to indices
    start_idx = cl('C')
    end_idx = cl('BJ')

    # Loop through columns in increments of 3 and merge cells
    for col in range(start_idx, end_idx + 1, 3):
        start = get_column_letter(col)
        end = get_column_letter(min(col + 2, end_idx))
        sheet.merge_cells(f'{start}8:{end}8')
        for i in range(cl(start), cl(end) + 1):
            sheet.cell(row=8, column=i).border = Border(top=Side(style='thin'),
                                                        bottom=Side(style='thin'),
                                                        left=Side(style='thin'),
                                                        right=Side(style='thin'))
        sheet[f"{start}8"].alignment = Alignment(vertical='center', horizontal='center')

    cols = ["OG", "Round", "AngleDev"]
    j = 0
    for col in cols:
        for i in range(start_idx + j, end_idx + j, 3):
            sheet.cell(row=9, column=i).value = col
            sheet.cell(row=9, column=i).alignment = Alignment(horizontal='center')
            sheet.cell(row=9, column=i).border = Border(top=Side(style='thin'),
                                                        bottom=Side(style='thin'),
                                                        right=Side(style='thin'))
        j += 1

    for col in range(2, 23):
        sheet.column_dimensions[chr(64 + col)].width = 9

    row = sheet.max_row + 1
    for index, df_row in df.iterrows():
        sheet.cell(row=row, column=2).value = df_row.name
        for i, val in enumerate(df_row, start=3):
            sheet.cell(row=row, column=i).value = val
        sheet.cell(row=row, column=2).border = Border(left=Side(style='thin'), right=Side(style='thin'))
        for i in range(3, 60 + 1, 3):
            sheet.cell(row=row, column=2 + i).border = Border(right=Side(style='thin'))
        for i in range(15, 60 + 1, 15):
            sheet.cell(row=row, column=2 + i).border = Border(right=Side(style='thick'))
        row += 1

    wb.save(filename=f"{filename}")
